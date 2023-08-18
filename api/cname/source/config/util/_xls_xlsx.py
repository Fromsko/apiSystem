# -*- coding: utf-8 -*-
"""
    @File  : xls_xlsx.py
    @Date  : 2023-03-17 21:58:54
    @notes : 文件格式转换 [xls => xlsx]
"""
import xlrd
from openpyxl.workbook import Workbook


def open_xls_as_xlsx(xls_path, xlsx_path):
    """将 Excel中 xls 样式改为 xlsx

    :param xls_path:  xls文件路径
    :param xlsx_path: xlsx文件保存路径
    :return:
    """
    # first open using xlrd
    book = xlrd.open_workbook(xls_path)
    index = 0
    nrows, ncols = 0, 0
    sheet = book.sheet_by_index(0)
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book_new = Workbook()
    sheet_new = book_new.create_sheet("sheet1", 0)

    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet_new.cell(row=row + 1, column=col +
                           1).value = sheet.cell_value(row, col)

    book_new.save(xlsx_path)
    return xlsx_path
